from __future__ import annotations

import hashlib
import html
import os
import shutil
import sys
import re
import tempfile
import traceback
from copy import copy
from io import BytesIO
from pathlib import Path
from urllib.parse import unquote, urlparse
from urllib.request import urlopen

import numpy as np
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


# -----------------------------------------------------------------------------
# 기본 경로 / 환경
# -----------------------------------------------------------------------------
APP_DIR = Path(__file__).resolve().parent
os.chdir(APP_DIR)

SKILL_SCRIPT = Path(__file__).resolve()

TEMP_ROOT = APP_DIR / ".eco_tmp"
TEMP_ROOT.mkdir(parents=True, exist_ok=True)

MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

I18N = {
    "ko": {
        "page_title": "BOM 친환경 체크리스트",
        "hero_title": "친환경 체크리스트 자동 생성기",
        "hero_desc": "BOM Excel을 업로드하면 친환경 체크리스트 워크북과 검토 항목을 한 번에 생성합니다.",
        "sidebar_title": "BOM 입력",
        "bom_upload": "BOM 파일 업로드",
        "auto_generate": "자동 생성",
        "sidebar_tip": "BOM만 업로드해도 바로 처리됩니다.",
        "language": "Language",
        "korean": "Korean",
        "english": "English",
        "model_name": "모델명",
        "plastic_total_weight": "플라스틱 총중량",
        "pcr_weight": "PCR 레진 중량",
        "pcr_ratio": "PCR 레진 비율",
        "part_list": "Part List",
        "plastic_part_list": "Plastic Part List",
        "metal_part_list": "Metal Part List",
        "download": "다운로드",
        "download_button": "📥 친환경 체크리스트 다운로드",
        "input_output": "입력 / 출력",
        "input_file": "입력 파일",
        "output_file": "출력 파일",
        "generated_message": "친환경 체크리스트 생성이 완료되었습니다.",
        "generated_caption": "생성된 워크북을 바로 내려받을 수 있습니다.",
        "input_file_wait": "왼쪽 사이드바에서 BOM 파일을 업로드하면, 결과가 이 화면에 자동으로 표시됩니다.",
        "support_title": "지원 기능",
        "support_1": "BOM 업로드 후 자동 생성",
        "support_2": "검토 필요 항목 자동 정리",
        "support_3": "완성 워크북 즉시 다운로드",
        "review_title": "검토 필요 항목",
        "review_details": "검토 상세 목록",
        "summary_title": "생성 결과",
        "plastic_rows": "Plastic rows",
        "metal_rows": "Metal rows",
        "review_rows": "검토 필요 항목",
        "rows": "rows",
        "count_suffix": "건",
    },
    "en": {
        "page_title": "BOM Eco Checklist",
        "hero_title": "Eco Checklist Generator",
        "hero_desc": "Upload a BOM Excel file to generate the eco checklist workbook and review items in one step.",
        "sidebar_title": "BOM Input",
        "bom_upload": "Upload BOM file",
        "auto_generate": "Auto generate",
        "sidebar_tip": "Upload only the BOM and it will process immediately.",
        "language": "Language",
        "korean": "Korean",
        "english": "English",
        "model_name": "Model",
        "plastic_total_weight": "Plastic total weight",
        "pcr_weight": "PCR resin weight",
        "pcr_ratio": "PCR resin ratio",
        "part_list": "Part List",
        "plastic_part_list": "Plastic Part List",
        "metal_part_list": "Metal Part List",
        "download": "Download",
        "download_button": "📥 Download Eco Checklist",
        "input_output": "Input / Output",
        "input_file": "Input file",
        "output_file": "Output file",
        "generated_message": "Eco checklist generation completed.",
        "generated_caption": "You can download the generated workbook immediately.",
        "input_file_wait": "Upload a BOM file from the sidebar and the result will appear here automatically.",
        "support_title": "Supported features",
        "support_1": "Auto generation after BOM upload",
        "support_2": "Automatic review item aggregation",
        "support_3": "Instant workbook download",
        "review_title": "Review items",
        "review_details": "Review details",
        "summary_title": "Generation result",
        "plastic_rows": "Plastic rows",
        "metal_rows": "Metal rows",
        "review_rows": "Review items",
        "rows": "rows",
        "count_suffix": "ea",
    },
}

GENERIC_MODEL_TOKENS = {
    "bom",
    "ebom",
    "mbom",
    "pbom",
    "tbom",
    "model",
    "template",
    "sample",
    "check",
    "checklist",
    "list",
    "eco",
    "eng",
    "download",
    "upload",
}

MODEL_LABELS = {
    "model",
    "modelname",
    "model no",
    "modelno",
    "model number",
    "modelnumber",
    "product model",
    "productmodel",
    "모델",
    "모델명",
}

PART_NO_LABELS = {
    "part no",
    "part no.",
    "part number",
    "p/no",
    "p-no",
    "품번",
}


# -----------------------------------------------------------------------------
# 유틸리티
# -----------------------------------------------------------------------------
def safe_stem(name: str) -> str:
    text = Path(name).stem if name else "upload"
    text = re.sub(r"[^\w가-힣.-]+", "_", text, flags=re.UNICODE)
    text = re.sub(r"_+", "_", text).strip("._ ")
    return text or "upload"


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def human_bytes(size: int) -> str:
    units = ["B", "KB", "MB", "GB"]
    value = float(size)
    for unit in units:
        if value < 1024 or unit == units[-1]:
            return f"{value:.1f} {unit}" if unit != "B" else f"{int(value)} B"
        value /= 1024
    return f"{size} B"


def tr(lang: str, key: str) -> str:
    return I18N.get(lang, I18N["ko"]).get(key, key)


def format_grams(value) -> str:
    try:
        if value is None:
            return "-"
        number = float(value)
        if number != number:
            return "-"
        return f"{number:,.1f} g"
    except Exception:
        return "-"


def format_percent(value) -> str:
    try:
        if value is None:
            return "-"
        number = float(value)
        if number != number:
            return "-"
        return f"{number:.1f}%"
    except Exception:
        return "-"


def save_uploaded_file(uploaded_file, prefix: str) -> Path:
    suffix = Path(uploaded_file.name).suffix or ".xlsx"
    digest = sha256_bytes(uploaded_file.getvalue())[:12]
    filename = f"{prefix}_{safe_stem(uploaded_file.name)}_{digest}{suffix}"
    path = TEMP_ROOT / filename
    path.write_bytes(uploaded_file.getvalue())
    return path


def normalize_model_candidate(text: str) -> str:
    value = re.sub(r"\s+", " ", str(text or "")).strip()
    if not value:
        return ""
    value = value.strip(" :;,_-()[]{}<>")
    value = re.sub(r"\.(xlsx|xlsm|xls|csv)$", "", value, flags=re.IGNORECASE)
    return value


def is_generic_model_token(text: str) -> bool:
    value = normalize_model_candidate(text).lower()
    if not value:
        return True
    compact = re.sub(r"[^a-z0-9가-힣]+", "", value.lower())
    return compact in GENERIC_MODEL_TOKENS


def split_filename_tokens(name: str) -> list[str]:
    stem = Path(name).stem if name else ""
    tokens = re.split(r"[\s_\-+(){}\[\]./\\]+", stem)
    return [token for token in tokens if token]


def score_model_token(token: str) -> int:
    text = normalize_model_candidate(token)
    if not text or is_generic_model_token(text):
        return -100

    compact = re.sub(r"[^A-Za-z0-9가-힣]+", "", text)
    if len(compact) < 4:
        return -50

    has_alpha = bool(re.search(r"[A-Za-z가-힣]", compact))
    has_digit = bool(re.search(r"\d", compact))

    score = len(compact)
    if has_alpha and has_digit:
        score += 25
    elif has_alpha:
        score += 8
    if re.search(r"[A-Za-z]{2,}\d{2,}", compact):
        score += 12
    if re.search(r"\d{2,}[A-Za-z]{1,}", compact):
        score += 6
    return score


def infer_model_name_from_filename(file_name: str) -> str:
    tokens = split_filename_tokens(file_name)
    if not tokens:
        return ""

    ranked = sorted(((score_model_token(token), token) for token in tokens), reverse=True)
    for score, token in ranked:
        if score > 0:
            return normalize_model_candidate(token)
    return ""


def cell_text(value) -> str:
    if value is None:
        return ""
    text = normalize_model_candidate(value)
    if text in {"None", "nan", "NaN"}:
        return ""
    return text


def infer_model_name_from_workbook(uploaded_file) -> str:
    try:
        wb = load_workbook(BytesIO(uploaded_file.getvalue()), data_only=True)
    except Exception:
        return ""

    try:
        # 0) Part No. / 품번 컬럼에서 첫 유효값을 먼저 찾는다.
        for ws in wb.worksheets[:3]:
            header_row_limit = min(ws.max_row or 0, 20)
            header_col_limit = min(ws.max_column or 0, 20)

            part_no_col = None
            part_no_header_row = None

            for row in ws.iter_rows(min_row=1, max_row=header_row_limit, min_col=1, max_col=header_col_limit):
                for cell in row:
                    text = cell_text(cell.value)
                    if not text:
                        continue
                    normalized_label = re.sub(r"[^a-z0-9가-힣]+", "", text.lower())
                    if normalized_label in {re.sub(r"[^a-z0-9가-힣]+", "", label.lower()) for label in PART_NO_LABELS}:
                        part_no_col = cell.column
                        part_no_header_row = cell.row
                        break
                if part_no_col is not None:
                    break

            if part_no_col is not None and part_no_header_row is not None:
                for row_idx in range(part_no_header_row + 1, min(ws.max_row or 0, part_no_header_row + 25) + 1):
                    raw_value = ws.cell(row=row_idx, column=part_no_col).value
                    candidate = normalize_model_candidate(raw_value)
                    if not candidate:
                        continue
                    candidate = re.split(r"[@\s]", candidate, maxsplit=1)[0]
                    candidate = re.split(r"[\.|/]", candidate, maxsplit=1)[0]
                    candidate = normalize_model_candidate(candidate)
                    if candidate and not is_generic_model_token(candidate):
                        return candidate

        # 1) 라벨 기반 추출
        for ws in wb.worksheets[:3]:
            max_row = min(ws.max_row or 0, 50)
            max_col = min(ws.max_column or 0, 20)
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    text = cell_text(cell.value)
                    if not text:
                        continue

                    normalized_label = re.sub(r"[^a-z0-9가-힣]+", "", text.lower())
                    if normalized_label in {re.sub(r"[^a-z0-9가-힣]+", "", label.lower()) for label in MODEL_LABELS}:
                        neighbors = []
                        if cell.column < ws.max_column:
                            neighbors.append(ws.cell(row=cell.row, column=cell.column + 1).value)
                            neighbors.append(ws.cell(row=cell.row, column=min(cell.column + 2, ws.max_column)).value)
                        if cell.row < ws.max_row:
                            neighbors.append(ws.cell(row=cell.row + 1, column=cell.column).value)

                        for neighbor in neighbors:
                            candidate = cell_text(neighbor)
                            if candidate and not is_generic_model_token(candidate):
                                return candidate

                    # "Model: ABC123" 형태
                    match = re.search(
                        r"(?:model(?:\s*name)?|모델명?)\s*[:\-]?\s*([A-Za-z0-9가-힣][A-Za-z0-9가-힣._\-\/ ]{2,})",
                        text,
                        re.IGNORECASE,
                    )
                    if match:
                        candidate = normalize_model_candidate(match.group(1))
                        if candidate and not is_generic_model_token(candidate):
                            return candidate

        # 2) 상단 영역의 강한 후보 추출
        candidates: list[tuple[int, str]] = []
        for ws in wb.worksheets[:3]:
            max_row = min(ws.max_row or 0, 30)
            max_col = min(ws.max_column or 0, 20)
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    text = cell_text(cell.value)
                    if not text or is_generic_model_token(text):
                        continue
                    compact = re.sub(r"[^A-Za-z0-9가-힣]+", "", text)
                    if len(compact) < 5:
                        continue
                    if not re.search(r"[A-Za-z가-힣]", compact):
                        continue
                    score = score_model_token(text)
                    if score > 0:
                        candidates.append((score, text))

        if candidates:
            candidates.sort(key=lambda item: item[0], reverse=True)
            return candidates[0][1]
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return ""


def infer_model_name(uploaded_file) -> str:
    if uploaded_file is None:
        return ""

    workbook_guess = infer_model_name_from_workbook(uploaded_file)
    if workbook_guess:
        return workbook_guess

    filename_guess = infer_model_name_from_filename(uploaded_file.name)
    if filename_guess:
        return filename_guess

    return ""


def build_signature(
    bom_file,
    resin_file,
    metal_file,
    template_file,
    header_row: int,
    sheet_name: str,
    model_name: str,
    output_dir: str,
    auto_generate: bool,
) -> str:
    parts: list[bytes] = []

    def add_file(uploaded):
        if uploaded is None:
            parts.append(b"<none>")
            return
        parts.append(uploaded.name.encode("utf-8", errors="ignore"))
        parts.append(uploaded.getvalue())

    add_file(bom_file)
    add_file(resin_file)
    add_file(metal_file)
    add_file(template_file)
    parts.append(str(header_row).encode("utf-8"))
    parts.append((sheet_name or "").encode("utf-8"))
    parts.append((model_name or "").encode("utf-8"))
    parts.append((output_dir or "").encode("utf-8"))

    digest = hashlib.sha256()
    for part in parts:
        digest.update(part)
        digest.update(b"\0")
    return digest.hexdigest()


# -----------------------------------------------------------------------------
# 내장 처리 로직 참조
# -----------------------------------------------------------------------------
def get_skill():
    return sys.modules[__name__]


# -----------------------------------------------------------------------------
# 생성 로직
# -----------------------------------------------------------------------------
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

MOLD_PREFIXES = ("MCK", "MBH", "MAZ", "MEY", "MAM")
PLASTIC_COLUMNS = [
    "P/N",
    "Plastic Component Name",
    "외관 노출부품 or 내부부품",
    "Weight (g)",
    "Type of Plastic",
    "Plastic Brand Name",
    "Plastic Model Name",
    "Applied Area",
    "Plastic coding Text",
    "Picture of Plastic coding",
    "Flame retardent type",
    "Halogen free check",
    "Paint & Spray Color&Method",
    "Mold-in or Glued-on metal",
    "Percent PCR plastic(%)",
    "Weight of PCR plastic(g)",
    "Remark",
    "규격확인 (OK/NG)",
]
METAL_COLUMNS = ["Part No.", "Part Name", "Metal Type", "thickness(mm)", "weight(g)"]
REVIEW_SHEET_NAME = "검토 필요 항목"
REVIEW_COLUMNS = [
    "구분",
    "BOM Row",
    "P/N",
    "Part Name / Component Name",
    "판정",
    "누락/미매칭 사유",
    "도면 확인 필요 항목",
    "권장 조치",
    "참고",
]

DEFAULT_RESIN_SOURCE = "http://hermes.lge.com/api/download/biz_archive/2452/Resin%20List.xlsx"
DEFAULT_METAL_SOURCE = "http://hermes.lge.com/api/download/biz_archive/2452/Metal%20List.xlsx"
DEFAULT_TEMPLATE_SOURCE = "http://hermes.lge.com/api/download/biz_archive/2452/MS%20%EC%A0%9C%ED%92%88%20%EC%B9%9C%ED%99%98%EA%B2%BD%20check%20list_ver%209.7_250430_%EA%B8%B0%EA%B5%AC.xlsx"

BOM_ALIASES = {
    "part_no": ["Part No", "Part No.", "P/No", "P-No", "Part Number", "Item No", "품번"],
    "part_name": ["Part Name", "Item Name", "부품명"],
    "technical_spec": ["Technical Spec", "Technical Specification", "Spec", "DETAIL"],
    "description": ["Description", "비고"],
    "qty": ["Qty", "QTY", "Quantity", "수량"],
    "material": ["Material", "재질"],
    "level": ["Level", "Lvl", "LVL", "Assembly Level", "층"],
    "category": ["Category", "Part Category", "부품군", "분류"],
    "remark": ["Remark", "Notes", "Comment", "비고"],
}

RESIN_ALIASES = {
    "Resin": ["Resin"],
    "Grade": ["Grade"],
    "Recycle Ratio": ["Recycle Ratio"],
    "Type of Plastic": ["Type of Plastic"],
    "Plastic Brand Name": ["Plastic Brand Name"],
    "Plastic coding Text": ["Plastic coding Text"],
    "Flame retardent type": ["Flame retardent type"],
    "Halogen free check": ["Halogen free check"],
}

METAL_ALIASES = {
    "P/No.": ["P/No.", "P/No", "Part No.", "Part No"],
    "Metal Type": ["Metal Type"],
    "Thickness(mm)": ["Thickness(mm)", "Thickness"],
    "Technical Specification": ["Technical Specification", "Technical Spec"],
}

METAL_PROCESS_TOKENS = {
    "CASTING",
    "CUTTING",
    "DIECASTING",
    "EXTRUSION",
    "FORGING",
    "PRESS",
}

METAL_EXCLUDE_TOKENS = {
    "ADHESIVE",
    "BOND",
    "COIL",
    "GLUE",
    "RAW MATERIAL",
}

METAL_MATERIAL_TOKENS = {
    "AL",
    "AL6063",
    "ALDC12",
    "EGI",
    "GI",
    "HGI",
    "PCM",
    "SCM",
    "SECC",
    "SGCC",
    "SGHC",
    "SPTE",
    "S50C",
    "STEEL",
    "SUS",
    "SUM24L",
}

METAL_THICKNESS_UNIT_RE = re.compile(r"\b(\d+(?:\.\d+)?)\s*T\b", re.IGNORECASE)
METAL_THICKNESS_MM_RE = re.compile(r"\b(\d+(?:\.\d+)?)\s*MM\b", re.IGNORECASE)
METAL_SPEC_TOKEN_RE = re.compile(r"\b([A-Z0-9]+)\b")
METAL_STANDALONE_NUMERIC_RE = re.compile(r"(?<![A-Z0-9/])(\d+(?:\.\d+)?)(?![A-Z0-9/])")


def canon(value):
    return re.sub(r"[\s\.\-_/()\[\]]+", "", str(value)).lower().strip()


def textify(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def source_to_bytes(source):
    if isinstance(source, (bytes, bytearray)):
        return bytes(source)

    source_text = str(source).strip()
    if source_text.lower().startswith(("http://", "https://")):
        with urlopen(source_text) as response:
            return response.read()

    return Path(source_text).expanduser().read_bytes()


def source_filename(source):
    source_text = str(source).strip()
    if source_text.lower().startswith(("http://", "https://")):
        parsed = urlparse(source_text)
        return unquote(Path(parsed.path).name)
    return Path(source_text).name


def compact_model_name(model_name):
    text = textify(model_name)
    if not text:
        return ""
    match = re.match(r"^[A-Za-z0-9]+", text)
    return match.group(0) if match else text


def make_clickable_filename(filename):
    stem = Path(filename).stem
    suffix = Path(filename).suffix
    clickable_stem = re.sub(r"\s+", "_", stem)
    clickable_stem = re.sub(r"_+", "_", clickable_stem).strip("_")
    return f"{clickable_stem}{suffix}"


def unique_file_path(directory, filename):
    directory = Path(directory)
    candidate = directory / filename
    if not candidate.exists():
        return candidate

    stem = candidate.stem
    suffix = candidate.suffix
    for idx in range(2, 1000):
        next_candidate = directory / f"{stem}_{idx}{suffix}"
        if not next_candidate.exists():
            return next_candidate
    raise FileExistsError(f"Could not find an available filename in {directory}")


def find_column(columns, aliases):
    alias_norm = {canon(item) for item in aliases}
    for col in columns:
        if canon(col) in alias_norm:
            return col
    return None


def rename_by_aliases(df, aliases_map):
    rename_map = {}
    for standard_name, aliases in aliases_map.items():
        found = find_column(df.columns, aliases)
        if found is not None:
            rename_map[found] = standard_name
    out = df.rename(columns=rename_map).copy()
    for standard_name in aliases_map.keys():
        if standard_name not in out.columns:
            out[standard_name] = ""
    return out


def load_bom(path, header_row=1, sheet_name=None):
    if sheet_name is None:
        sheet_name = 0
    df = pd.read_excel(path, sheet_name=sheet_name, header=header_row - 1, engine="openpyxl")
    df = rename_by_aliases(df, BOM_ALIASES)
    df["qty"] = pd.to_numeric(df["qty"], errors="coerce")
    return df


def is_substitute_level(row):
    level = textify(row.get("level")).replace(" ", "").upper()
    return bool(level) and re.fullmatch(r"\*S\*", level) is not None


def row_level(row):
    level_text = textify(row.get("level")).replace(" ", "")
    if not level_text:
        return None
    match = re.search(r"-?\d+", level_text)
    if match is None:
        return None
    try:
        return int(match.group(0))
    except ValueError:
        return None


def should_stop_plastic_scan(current_level, sub_level):
    if current_level is None or sub_level is None:
        return False
    # Plastic 후보의 하위 원재료는 반드시 더 깊은 레벨에만 있어야 한다.
    # 같은 레벨은 형제 부품이므로, 다음 후보부터는 더 이상 현재 부품의 하위로 보지 않는다.
    return sub_level <= current_level


def load_resin_reference(source):
    df = pd.read_excel(BytesIO(source_to_bytes(source)), header=1, engine="openpyxl")
    df = rename_by_aliases(df, RESIN_ALIASES)
    df = df[list(RESIN_ALIASES.keys())].copy()
    df["Recycle Ratio"] = pd.to_numeric(df["Recycle Ratio"], errors="coerce").fillna(0)
    return df


def load_metal_reference(source):
    df = pd.read_excel(BytesIO(source_to_bytes(source)), engine="openpyxl")
    df = rename_by_aliases(df, METAL_ALIASES)
    return df[list(METAL_ALIASES.keys())].copy()


def make_review_item(section, bom_row, part_no, part_name, status, reason, check_item, action, reference=""):
    return {
        "구분": section,
        "BOM Row": bom_row,
        "P/N": part_no,
        "Part Name / Component Name": part_name,
        "판정": status,
        "누락/미매칭 사유": reason,
        "도면 확인 필요 항목": check_item,
        "권장 조치": action,
        "참고": reference,
    }


def lookup_exact_or_prefix(df, key_col, code):
    target = canon(code)
    if not target:
        return None

    exact_matches = df[df[key_col].map(lambda v: canon(v) == target)]
    if not exact_matches.empty:
        return exact_matches.iloc[0]

    candidates = []
    for _, row in df.iterrows():
        ref_code = canon(row.get(key_col, ""))
        if not ref_code:
            continue
        if target.startswith(ref_code) or ref_code.startswith(target):
            candidates.append((len(ref_code), row))
    if candidates:
        candidates.sort(key=lambda item: item[0], reverse=True)
        return candidates[0][1]
    return None


def find_resin_match(df, code):
    return lookup_exact_or_prefix(df, "Resin", code)


def find_metal_match(df, code):
    return lookup_exact_or_prefix(df, "P/No.", code)


def metal_source_rank(source):
    return 0 if source == "metal_list" else 1


def _metal_spec_text(row):
    return textify(row.get("technical_spec")) or primary_name(row) or textify(row.get("description"))


def _normalize_al_metal_type(text):
    text = textify(text).upper()
    if not text:
        return ""

    match = re.search(r"\bAL\s*([0-9]{3,4})\b", text, re.IGNORECASE)
    if match is not None:
        return f"AL{match.group(1)}"

    match = re.search(r"\bAL([0-9]{3,4})\b", text, re.IGNORECASE)
    if match is not None:
        return f"AL{match.group(1)}"

    if re.search(r"\bALDC12\b", text, re.IGNORECASE):
        return "ALDC12"
    if re.search(r"\bAL6063\b", text, re.IGNORECASE):
        return "AL6063"

    return ""


def _metal_type_pattern(metal_type):
    metal_type = textify(metal_type).upper()
    if not metal_type:
        return ""

    if metal_type == "AL":
        return r"\bAL\b"

    match = re.fullmatch(r"AL([0-9]{3,4})", metal_type)
    if match is not None:
        return rf"\bAL\s*{match.group(1)}\b"

    return rf"\b{re.escape(metal_type)}\b"


def _is_plausible_metal_thickness(value):
    """
    Thickness should be a realistic numeric value, not an alloy code or model token.
    Examples to reject: 6063, 5052, 55C6, 16A.
    """
    value = textify(value).strip()
    if not value:
        return False
    try:
        numeric = float(value)
    except Exception:
        return False
    # Bare numeric thickness values in this workflow are expected to be small.
    # Reject alloy-like or model-like numbers that are clearly not thickness.
    if numeric <= 0 or numeric >= 100:
        return False
    return True


def _find_standalone_numeric(text, first=True):
    matches = list(METAL_STANDALONE_NUMERIC_RE.finditer(textify(text).upper()))
    if not matches:
        return ""
    match = matches[0] if first else matches[-1]
    return match.group(1)


def _spec_has_token(spec_text, token):
    if not spec_text or not token:
        return False
    return re.search(rf"\b{re.escape(token)}\b", spec_text, re.IGNORECASE) is not None


def extract_metal_type_from_spec(spec_text):
    text = textify(spec_text).upper()
    if not text:
        return ""

    al_alloy = _normalize_al_metal_type(text)
    if al_alloy:
        return al_alloy

    # 우선순위가 높은 금속 재질명부터 찾는다.
    for token in ("ALDC12", "AL6063", "EGI", "HGI", "PCM", "SUS", "SPTE", "SECC", "SGCC", "SGHC", "S50C", "SUM24L", "SCM", "STEEL", "AL"):
        if _spec_has_token(text, token):
            return token

    # 공정명이 명확한 경우의 기본 재질값
    if "CASTING" in text or "DIECASTING" in text:
        return "ALDC12"
    if "EXTRUSION AL" in text:
        return "AL6063"

    return ""


def is_casting_or_diecasting_spec(spec_text):
    text = textify(spec_text).upper()
    if not text:
        return False
    return bool(re.search(r"\b(CASTING|DIECASTING)\b", text))


def extract_metal_thickness_from_spec(spec_text, metal_type=""):
    text = textify(spec_text).upper()
    if not text:
        return ""

    # CASTING / DIECASTING 계열은 두께 표기 대상이 아니므로 어떤 숫자도 thickness로 해석하지 않는다.
    if is_casting_or_diecasting_spec(text):
        return ""

    # EXTRUSION(압출) 계열은 단면 치수/외형 치수가 함께 들어오는 경우가 많아
    # thickness(mm)로 해석하지 않는다.
    if "EXTRUSION" in text:
        return ""

    # 가장 안전한 형태: 1.5T, 2.0T처럼 단위가 명시된 경우
    match = METAL_THICKNESS_UNIT_RE.search(text)
    if match is not None:
        return match.group(1)

    # 2.0mm처럼 단위가 mm로 명시된 경우
    match = METAL_THICKNESS_MM_RE.search(text)
    if match is not None:
        return match.group(1)

    process_match = None
    for token in ("PRESS", "CUTTING", "FORGING", "EXTRUSION", "CASTING", "DIECASTING"):
        process_match = re.search(rf"\b{re.escape(token)}\b", text, re.IGNORECASE)
        if process_match is not None:
            break

    metal_match = None
    if metal_type:
        pattern = _metal_type_pattern(metal_type)
        if pattern:
            metal_match = re.search(pattern, text, re.IGNORECASE)

    # 범용 규칙:
    # 1) 공정명과 재질명 사이에 있는 독립 숫자를 두께로 우선 판정한다.
    # 2) 재질명 뒤에 붙는 독립 숫자를 그 다음 후보로 본다.
    # 3) 공정명만 있는 경우 공정명 뒤의 독립 숫자를 본다.
    if process_match is not None and metal_match is not None:
        between = text[process_match.end():metal_match.start()]
        candidate = _find_standalone_numeric(between, first=False)
        if _is_plausible_metal_thickness(candidate):
            return candidate

        after = text[metal_match.end():]
        candidate = _find_standalone_numeric(after, first=True)
        if _is_plausible_metal_thickness(candidate):
            return candidate

    if metal_match is not None:
        before = text[:metal_match.start()]
        candidate = _find_standalone_numeric(before, first=False)
        if _is_plausible_metal_thickness(candidate):
            return candidate

        after = text[metal_match.end():]
        candidate = _find_standalone_numeric(after, first=True)
        if _is_plausible_metal_thickness(candidate):
            return candidate

    if process_match is not None:
        after = text[process_match.end():]
        candidate = _find_standalone_numeric(after, first=True)
        if _is_plausible_metal_thickness(candidate):
            return candidate

    # 명시적 공정/재질 위치가 없더라도, 독립 숫자만 마지막으로 제한적으로 허용
    candidate = _find_standalone_numeric(text, first=True)
    if _is_plausible_metal_thickness(candidate):
        return candidate

    return ""


MATERIAL_INFO_PATTERNS = (
    r"\bRESIN\b",
    r"\bMATERIAL\b",
    r"\bABS\b",
    r"\bPOM\b",
    r"\bPA\d*\b",
    r"\bPBT\b",
    r"\bHIPS\b",
    r"\bSAN\b",
    r"\bPMMA\b",
    r"\bPVC\b",
    r"\bPPE\b",
    r"\bPPS\b",
    r"\bPCTG\b",
    r"\bTPE\b",
    r"\bTPU\b",
    r"\bPU\b",
)

PLASTIC_COLOR_TOKENS = {
    "BK",
    "BLK",
    "BLACK",
    "BL",
    "BLUE",
    "BR",
    "BROWN",
    "GD",
    "GOLD",
    "GF",
    "GR",
    "GRAY",
    "GREY",
    "GN",
    "GREEN",
    "OR",
    "ORANGE",
    "PK",
    "PINK",
    "PR",
    "PURPLE",
    "PURBLE",
    "PURPLEGRAY",
    "PURBLEGRAY",
    "RD",
    "RED",
    "SL",
    "SILVER",
    "WH",
    "WHITE",
    "VI",
    "VIOLET",
}

LEADING_PLASTIC_TOKENS = {
    "ABS",
    "HIPS",
    "MOLD",
    "PC",
    "PCTG",
    "PE",
    "PBT",
    "PCGF",
    "PEEK",
    "POM",
    "PP",
    "PPE",
    "PPS",
    "PS",
    "PVC",
    "PU",
    "PA",
    "PA6",
    "PA66",
    "PA12",
    "RESIN",
    "SAN",
    "TPE",
    "TPU",
}

LEADING_METAL_TOKENS = {
    "AL",
    "AL6063",
    "ALDC12",
    "CASTING",
    "EGI",
    "HGI",
    "MOLD",
    "N",
    "PRESS",
    "SECC",
    "SGCC",
    "SGHC",
}

COMMON_AMBIGUOUS_TOKENS = {
    "ASSY",
    "ASSEMBLY",
    "BOARD",
    "MAIN",
    "MODULE",
    "PCB",
    "SMPS",
    "UNIT",
}

MEANINGFUL_SHORT_TOKENS = {
    "ARM",
    "BACK",
    "BASE",
    "BODY",
    "BOTTOM",
    "BRACKET",
    "BUTTON",
    "CABINET",
    "CASE",
    "CLIP",
    "COVER",
    "FRAME",
    "FRONT",
    "GUIDE",
    "HOLDER",
    "HINGE",
    "HOOK",
    "JOY",
    "KNOB",
    "LATCH",
    "LCD",
    "LED",
    "LID",
    "LEFT",
    "LOCK",
    "LOW",
    "LOWER",
    "OUTER",
    "PAD",
    "PANEL",
    "PLATE",
    "PIVOT",
    "RAIL",
    "REAR",
    "RIGHT",
    "SHIELD",
    "SIDE",
    "SPRING",
    "STAND",
    "SUPPORT",
    "TAB",
    "TILT",
    "TOP",
    "TRAY",
    "TCON",
    "T-CON",
    "UP",
    "UPPER",
    "USB",
    "VESA",
    "WHEEL",
}

CORE_PART_TOKENS = {
    "ARM",
    "BACK",
    "BASE",
    "BODY",
    "BOTTOM",
    "BRACKET",
    "BUTTON",
    "CABINET",
    "CASE",
    "CLIP",
    "COVER",
    "CONTROL",
    "FIX",
    "FRAME",
    "FRONT",
    "GUIDE",
    "HOLDER",
    "HINGE",
    "HOOK",
    "JOY",
    "KNOB",
    "LATCH",
    "LOCK",
    "LID",
    "OUTER",
    "PAD",
    "PANEL",
    "PLATE",
    "PIVOT",
    "RAIL",
    "REAR",
    "RIGHT",
    "SHIELD",
    "SIDE",
    "SPRING",
    "STAND",
    "SUPPORT",
    "TAB",
    "TILT",
    "TOP",
    "TRAY",
    "SPEAKER",
    "VESA",
    "WHEEL",
}

NOTE_TOKENS = {
    "COST",
    "FOR",
    "HEIGHT",
    "LOW",
    "MODEL",
    "ONE",
    "NT",
    "CLICK",
}

WEAK_ONLY_TOKENS = {
    "BOTTOM",
    "FRONT",
    "INNER",
    "LEFT",
    "LOW",
    "LOWER",
    "MAIN",
    "OUTER",
    "RIGHT",
    "SIDE",
    "TOP",
    "UP",
    "UPPER",
}

COMPONENT_TRAILING_NOISE_TOKENS = {
    "ABS",
    "ASA",
    "CHEMICAL",
    "CHEMICALS",
    "CO",
    "COMPANY",
    "CORP",
    "CORPORATION",
    "ELECTRONICS",
    "INDUSTRIAL",
    "KOREA",
    "LG",
    "LOTTE",
    "LTD",
    "LTD.",
    "MATERIAL",
    "MATERIALS",
    "PC",
    "PC+ABS",
    "PC-ABS",
    "PA",
    "PA12",
    "PA66",
    "PA6",
    "PBT",
    "PE",
    "PEEK",
    "POM",
    "PP",
    "PPE",
    "PPS",
    "PCTG",
    "PMMA",
    "PVC",
    "RESIN",
    "SAMSUNG",
    "SAN",
    "TECH",
    "TECHNOLOGY",
    "TPE",
    "TPU",
}


def row_has_material_info(row):
    if textify(row.get("material")):
        return True

    haystack = " ".join(
        part for part in (
            textify(row.get("part_name")),
            textify(row.get("technical_spec")),
            textify(row.get("description")),
        )
        if part
    ).upper()
    if not haystack:
        return False

    return any(re.search(pattern, haystack) for pattern in MATERIAL_INFO_PATTERNS)


def primary_name(row):
    for key in ("technical_spec", "part_name", "description"):
        value = textify(row.get(key))
        if value:
            return value
    return ""


def combined_component_text(row):
    return " ".join(
        part for part in (
            textify(row.get("part_name")),
            textify(row.get("technical_spec")),
            textify(row.get("description")),
        )
        if part
    ).strip()


def _normalize_component_tokens(text):
    normalized = textify(text).replace("\u00a0", " ")
    normalized = normalized.replace("/", " ")
    normalized = re.sub(r"\s+", " ", normalized).strip()
    tokens = []
    for token in normalized.split(" "):
        token = token.strip().strip("()[]{}")
        token = token.strip(".,:;")
        if token:
            tokens.append(token)
    return tokens


def _token_raw(token):
    return textify(token).strip("()[]{}.,:;")


def _token_upper(token):
    return _token_raw(token).upper()


def _token_compact(token):
    return re.sub(r"[_\-\s]+", "", _token_upper(token))


def _token_parts(token):
    raw = _token_upper(token)
    return [part for part in re.split(r"[_]+", raw) if part]


def _token_has_core(token):
    raw = _token_upper(token)
    if raw in CORE_PART_TOKENS or raw in MEANINGFUL_SHORT_TOKENS:
        return True
    if _token_compact(token) in CORE_PART_TOKENS:
        return True
    return any(part in CORE_PART_TOKENS or part in MEANINGFUL_SHORT_TOKENS for part in _token_parts(token))


def _looks_like_code_token(token):
    raw = _token_upper(token)
    if not raw:
        return False

    if raw in LEADING_PLASTIC_TOKENS or raw in LEADING_METAL_TOKENS:
        return True
    if raw in PLASTIC_COLOR_TOKENS:
        return True
    compact = _token_compact(token)
    if compact in LEADING_PLASTIC_TOKENS or compact in LEADING_METAL_TOKENS:
        return True
    if compact in PLASTIC_COLOR_TOKENS:
        return True
    if re.fullmatch(r"(?:PC|PBT|POM|ABS|HIPS|PP|PE|PA\d{0,2}|PPS|SAN|PMMA|PVC|PCTG|TPE|TPU|PU)(?:\+GF\d+%?)?", raw):
        return True
    if re.fullmatch(r"(?:EGI|HGI|SECC|SGCC|SGHC|AL|ALDC12|AL6063)", raw):
        return True
    if re.fullmatch(r"\d+(?:\.\d+)?T?", raw):
        return True
    if re.fullmatch(r"\d{2,}[A-Z0-9\-]*", compact):
        return True
    if re.fullmatch(r"[A-Z]{1,4}\d[A-Z0-9\-]*", compact):
        return True
    if re.fullmatch(r"[A-Z0-9]+(?:\+[A-Z0-9]+)+", raw):
        return True
    return False


def _is_trailing_noise_token(token):
    raw = _token_upper(token)
    if not raw:
        return True
    if raw in PLASTIC_COLOR_TOKENS:
        return True
    compact = _token_compact(token)
    if compact in PLASTIC_COLOR_TOKENS:
        return True
    if raw in COMPONENT_TRAILING_NOISE_TOKENS or compact in COMPONENT_TRAILING_NOISE_TOKENS:
        return True
    if raw in {"NONE", "NT_TOOL", "TOOL"}:
        return True
    if compact in {"NONE", "N","TOOL"}:
        return True
    if re.fullmatch(r"\d{1,2}-[A-Z]{1,3}", raw):
        return True
    return False


def _is_color_token(token):
    raw = _token_upper(token)
    if not raw:
        return False
    compact = _token_compact(token)
    return raw in PLASTIC_COLOR_TOKENS or compact in PLASTIC_COLOR_TOKENS or raw in {"NONE", "PURBLE", "PURPLE"}


def _clean_component_token(token, kind):
    raw = _token_raw(token)
    if not raw:
        return ""

    if "_" not in raw:
        if kind == "plastic" and raw.upper() == "FIBER":
            return ""
        if _looks_like_code_token(raw) or _is_trailing_noise_token(raw):
            return ""
        return raw

    parts = [part for part in raw.split("_") if part]
    if not parts:
        return ""

    while parts:
        head = parts[0]
        if kind == "plastic":
            if head.upper() == "MOLD" or _looks_like_code_token(head) or head.upper() == "FIBER":
                parts.pop(0)
                continue
        else:
            if head.upper() in LEADING_METAL_TOKENS or _looks_like_code_token(head):
                parts.pop(0)
                continue
        break

    while parts and _is_color_token(parts[-1]):
        parts.pop()

    if not parts:
        return ""

    if len(parts) == 1:
        only = parts[0]
        if _is_trailing_noise_token(only):
            return ""
        if only.upper() not in CORE_PART_TOKENS and only.upper() not in MEANINGFUL_SHORT_TOKENS:
            if only.isalpha() and only.upper() == only and len(only) <= 4:
                return ""

    return "_".join(parts)


def _strip_component_noise(tokens, kind):
    cleaned = list(tokens)

    while cleaned:
        head = _token_upper(cleaned[0])
        if kind == "plastic":
            if head == "MOLD" or _looks_like_code_token(head):
                cleaned.pop(0)
                continue
            if head == "FIBER":
                cleaned.pop(0)
                continue
        else:
            if head in LEADING_METAL_TOKENS or _looks_like_code_token(head):
                cleaned.pop(0)
                continue
        break

    while cleaned and _is_trailing_noise_token(cleaned[-1]):
        cleaned.pop()

    return cleaned


def extract_component_candidate(raw_text, kind):
    text = textify(raw_text)
    if not text:
        return ""

    segments = []
    split_text = re.split(r"\s*[,;]\s*|\s+-\s+", text)
    for piece in split_text:
        piece_tokens = _normalize_component_tokens(piece)
        if not piece_tokens:
            continue

        current = []
        for token in piece_tokens:
            cleaned = _clean_component_token(token, kind)
            if not cleaned:
                if current:
                    segments.append(current)
                    current = []
                continue
            current.append(cleaned)
        if current:
            segments.append(current)

    if not segments:
        return ""

    def segment_score(segment):
        upper_tokens = [_token_upper(token) for token in segment]
        core_count = sum(1 for token in upper_tokens if _token_has_core(token))
        return (core_count, len(segment), len(" ".join(segment)))

    normalized_segments = []
    for segment in segments:
        stripped = _strip_component_noise(segment, kind)
        if stripped:
            normalized_segments.append(stripped)
        else:
            normalized_segments.append(segment)

    usable_segments = [segment for segment in normalized_segments if is_component_name_usable(" ".join(segment).strip())]
    if usable_segments:
        usable_segments.sort(key=segment_score, reverse=True)
        return " ".join(usable_segments[0]).strip()

    return ""


def extract_tail_component_candidate(raw_text, kind):
    text = textify(raw_text)
    if not text:
        return ""

    tail_parts = [part.strip() for part in re.split(r"\s*[,;]\s*", text) if part.strip()]
    if len(tail_parts) <= 1:
        return ""

    tail_candidate = extract_component_candidate(tail_parts[-1], kind)
    if is_component_name_usable(tail_candidate):
        return tail_candidate
    return ""


def is_component_name_usable(candidate):
    candidate = textify(candidate)
    if not candidate:
        return False

    tokens = [token for token in candidate.split() if token]
    if not tokens:
        return False

    upper_tokens = [_token_upper(token) for token in tokens]
    has_core_token = any(_token_has_core(token) for token in tokens)

    if not has_core_token:
        return False

    if len(upper_tokens) == 1 and upper_tokens[0] in WEAK_ONLY_TOKENS:
        return False

    if all(_looks_like_code_token(token) or _is_trailing_noise_token(token) for token in upper_tokens):
        return False

    if len(upper_tokens) >= 3 and any(token in COMMON_AMBIGUOUS_TOKENS for token in upper_tokens):
        return False

    return True


def resolve_component_name(row, kind):
    if kind == "plastic":
        speaker_candidate = extract_speaker_component_name(row)
        if is_component_name_usable(speaker_candidate):
            return speaker_candidate

    for key in ("part_name", "technical_spec", "description"):
        if kind == "plastic":
            tail_candidate = extract_tail_component_candidate(row.get(key), kind)
            if is_component_name_usable(tail_candidate):
                return tail_candidate
        candidate = extract_component_candidate(row.get(key), kind)
        if is_component_name_usable(candidate):
            return candidate

    fallback = primary_name(row)
    if kind == "plastic":
        tail_candidate = extract_tail_component_candidate(fallback, kind)
        if is_component_name_usable(tail_candidate):
            return tail_candidate
    candidate = extract_component_candidate(fallback, kind)
    if is_component_name_usable(candidate):
        return candidate
    return fallback


def extract_speaker_component_name(row):
    haystack = combined_component_text(row)
    if not haystack or "SPEAKER" not in haystack.upper():
        return ""

    side = ""
    side_patterns = (
        (r"\bLEFT\b", "LEFT"),
        (r"\bRIGHT\b", "RIGHT"),
        (r"\bL\s*/\s*R\b", "L/R"),
    )
    for pattern, label in side_patterns:
        if re.search(pattern, haystack, re.IGNORECASE):
            side = label
            break

    candidate = "Speaker"
    if side:
        candidate = f"Speaker {side}"
    return candidate


def is_plastic_candidate(row):
    part_no = textify(row.get("part_no")).upper()
    spec = primary_name(row).upper()
    combined = combined_component_text(row).upper()
    if not part_no:
        return False
    if "SPEAKER" in combined:
        return True
    return ("MOLD" in spec) and any(part_no.startswith(prefix) for prefix in MOLD_PREFIXES)


def is_raw_material_like_metal_row(row):
    part_no = textify(row.get("part_no")).upper()
    part_name = textify(row.get("part_name")).upper()
    spec = primary_name(row).upper()
    description = textify(row.get("description")).upper()
    haystack = " ".join(part for part in (part_no, part_name, spec, description) if part).upper()
    if not haystack:
        return False

    if any(token in haystack for token in METAL_EXCLUDE_TOKENS):
        return True

    # RCL 계열은 대체로 원재료/원단위 자재이므로, Coil/Raw Material 성격이면 Metal Part에서 제외한다.
    if "RCL" in part_no and any(token in haystack for token in ("COIL", "ROLL", "RAW MATERIAL")):
        return True

    return False


def should_stop_metal_scan(current_level, sub_level):
    if current_level is None or sub_level is None:
        return False
    return sub_level <= current_level


def is_metal_raw_material_row(row):
    part_no = textify(row.get("part_no")).upper()
    if "RCL" in part_no:
        return True
    return is_raw_material_like_metal_row(row)


def collect_metal_child_raw_materials(bom_df, parent_idx):
    parent_row = bom_df.iloc[parent_idx]
    current_level = row_level(parent_row)
    child_rows = []

    for sub_idx in range(parent_idx + 1, len(bom_df)):
        sub_row = bom_df.iloc[sub_idx]
        if is_substitute_level(sub_row):
            continue

        sub_level = row_level(sub_row)
        if should_stop_metal_scan(current_level, sub_level):
            break

        if is_metal_raw_material_row(sub_row):
            child_rows.append((sub_idx, sub_row))

    return child_rows


def is_metal_candidate(row):
    part_no = textify(row.get("part_no")).upper()
    spec = primary_name(row).upper()
    if not part_no or not spec:
        return False

    if is_raw_material_like_metal_row(row):
        return False

    if "CASTING" in spec or "EXTRUSION AL" in spec:
        return True

    has_process_token = any(token in spec for token in METAL_PROCESS_TOKENS)
    has_material_token = any(_spec_has_token(spec, token) for token in METAL_MATERIAL_TOKENS)
    return has_process_token and has_material_token


def extract_plastic_records(bom_df, resin_df, review_items=None):
    records = []
    total_weight = 0.0
    total_pcr_weight = 0.0

    for idx, row in bom_df.iterrows():
        if is_substitute_level(row):
            continue
        if not is_plastic_candidate(row):
            continue

        part_no = textify(row.get("part_no"))
        component_name = resolve_component_name(row, "plastic")
        current_level = row_level(row)
        weight = np.nan
        resin_match = None
        resin_part_no = ""
        has_lower_material_info = False

        for sub_idx in range(idx + 1, len(bom_df)):
            sub_row = bom_df.iloc[sub_idx]
            if is_substitute_level(sub_row):
                continue
            sub_level = row_level(sub_row)
            if should_stop_plastic_scan(current_level, sub_level):
                break
            sub_spec = primary_name(sub_row).upper()
            sub_part_no = textify(sub_row.get("part_no"))
            if current_level is None and ("MOLD" in sub_spec or "PRESS" in sub_spec):
                break
            if row_has_material_info(sub_row):
                has_lower_material_info = True
            if not sub_part_no:
                continue
            match = find_resin_match(resin_df, sub_part_no)
            if match is not None:
                resin_match = match
                resin_part_no = sub_part_no
                if pd.notna(sub_row.get("qty")):
                    weight = float(sub_row.get("qty")) * 1000
                break

        if pd.notna(weight) and weight < 0.5:
            continue

        if resin_match is not None:
            recycle_ratio = float(resin_match.get("Recycle Ratio", 0) or 0)
            percent_pcr = recycle_ratio * 100
            pcr_weight = weight * percent_pcr / 100 if pd.notna(weight) else np.nan
            if pd.notna(weight):
                total_weight += float(weight)
            if pd.notna(pcr_weight):
                total_pcr_weight += float(pcr_weight)
            record = {
                "P/N": part_no,
                "Plastic Component Name": component_name,
                "외관 노출부품 or 내부부품": "",
                "Weight (g)": weight,
                "Type of Plastic": textify(resin_match.get("Type of Plastic")),
                "Plastic Brand Name": textify(resin_match.get("Plastic Brand Name")),
                "Plastic Model Name": textify(resin_match.get("Grade")),
                "Applied Area": "W/W",
                "Plastic coding Text": textify(resin_match.get("Plastic coding Text")),
                "Picture of Plastic coding": "",
                "Flame retardent type": textify(resin_match.get("Flame retardent type")),
                "Halogen free check": textify(resin_match.get("Halogen free check")),
                "Paint & Spray Color&Method": "",
                "Mold-in or Glued-on metal": "",
                "Percent PCR plastic(%)": percent_pcr,
                "Weight of PCR plastic(g)": pcr_weight,
                "Remark": "",
                "규격확인 (OK/NG)": "",
            }
        else:
            material_hint = textify(row.get("material"))
            spec_hint = textify(row.get("technical_spec"))
            desc_hint = textify(row.get("description"))
            if has_lower_material_info:
                reason = "BOM에 재질 정보는 있으나 Resin List에 해당 원자재 정보가 없음"
                check_item = "BOM에 적힌 재질 코드, 사양 표기, 후속 resin 코드"
                action = "누락된 항목은 도면 확인하여 사용자가 직접 입력해주세요. (관리자에게 Resin List를 최신정보로 업데이트 요청)"
                reference = f"Spec={spec_hint or '-'} | Desc={desc_hint or '-'} | Material={material_hint or '-'}"
            else:
                reason = "BOM에 재질 정보가 없음."
                check_item = "도면의 재질 표기, BOM 하위 재질 행, 사양서"
                action = "누락된 항목은 도면 확인하여 사용자가 직접 입력해주세요."
                reference = ""

            if review_items is not None:
                review_items.append(
                    make_review_item(
                        section="Plastic Parts",
                        bom_row=idx + 2,
                        part_no=part_no,
                        part_name=component_name,
                        status="Review",
                        reason=reason,
                        check_item=check_item,
                        action=action,
                        reference=reference,
                    )
                )

            record = {
                "P/N": part_no,
                "Plastic Component Name": component_name,
                "외관 노출부품 or 내부부품": "",
                "Weight (g)": weight,
                "Type of Plastic": "",
                "Plastic Brand Name": "",
                "Plastic Model Name": "",
                "Applied Area": "W/W",
                "Plastic coding Text": "",
                "Picture of Plastic coding": "",
                "Flame retardent type": "",
                "Halogen free check": "",
                "Paint & Spray Color&Method": "",
                "Mold-in or Glued-on metal": "",
                "Percent PCR plastic(%)": np.nan,
                "Weight of PCR plastic(g)": np.nan,
                "Remark": "",
                "규격확인 (OK/NG)": "",
            }

        records.append(record)

    return pd.DataFrame(records, columns=PLASTIC_COLUMNS), total_weight, total_pcr_weight


def extract_metal_records(bom_df, metal_df, review_items=None):
    candidate_rows = []

    for idx, row in bom_df.iterrows():
        if is_substitute_level(row):
            continue
        if not is_metal_candidate(row):
            continue

        part_no = textify(row.get("part_no"))
        spec_text = _metal_spec_text(row)
        part_name = resolve_component_name(row, "metal")
        metal_type = extract_metal_type_from_spec(spec_text)
        thickness = extract_metal_thickness_from_spec(spec_text, metal_type=metal_type)
        thickness_lock = is_casting_or_diecasting_spec(spec_text)
        weight = np.nan
        source = "technical_spec"
        parent_match = find_metal_match(metal_df, part_no)
        if parent_match is not None:
            list_metal_type = textify(parent_match.get("Metal Type"))
            list_thickness = textify(parent_match.get("Thickness(mm)"))
            if list_metal_type or list_thickness:
                metal_type = list_metal_type or metal_type
                if not thickness_lock:
                    thickness = list_thickness or thickness
                source = "metal_list"

        review_entries = []
        child_raw_materials = collect_metal_child_raw_materials(bom_df, idx)
        child_part_nos = []

        if child_raw_materials:
            child_weight_total = 0.0
            child_weight_found = False
            child_missing_qty = False
            child_missing_list = False
            child_match_found = False

            for child_idx, child_row in child_raw_materials:
                child_part_no = textify(child_row.get("part_no"))
                child_part_nos.append(child_part_no)
                child_spec_text = _metal_spec_text(child_row)
                child_thickness_lock = is_casting_or_diecasting_spec(child_spec_text)
                child_match = find_metal_match(metal_df, child_part_no)
                child_qty = pd.to_numeric(child_row.get("qty"), errors="coerce")

                if pd.notna(child_qty):
                    child_weight_total += float(child_qty) * 1000
                    child_weight_found = True
                else:
                    child_missing_qty = True

                if child_match is not None:
                    list_metal_type = textify(child_match.get("Metal Type"))
                    list_thickness = textify(child_match.get("Thickness(mm)"))
                    if list_metal_type or list_thickness:
                        metal_type = list_metal_type or metal_type
                        if not child_thickness_lock and not thickness_lock:
                            thickness = list_thickness or thickness
                        source = "metal_list"
                        child_match_found = True
                else:
                    child_missing_list = True
                    if not metal_type:
                        metal_type = extract_metal_type_from_spec(child_spec_text) or metal_type
                    if not thickness and not child_thickness_lock and not thickness_lock:
                        thickness = extract_metal_thickness_from_spec(child_spec_text, metal_type=metal_type) or thickness

            if child_weight_found and not child_missing_qty:
                weight = child_weight_total

            if child_missing_list:
                review_entries.append(
                    make_review_item(
                        section="Metal Parts",
                        bom_row=idx + 2,
                        part_no=part_no,
                        part_name=part_name,
                        status="Review",
                        reason="하위 원재료는 있으나 Metal List에 해당 원재료 정보가 없음",
                        check_item="하위 BOM 원재료 코드, Metal List 등록 여부, 재질/두께 표기",
                        action="Metal List를 최신 정보로 업데이트하거나 도면에서 원재료 재질/두께를 확인해 주세요",
                        reference=f"Child raw material={', '.join(child_part_nos)}",
                    )
                )

            if child_missing_qty:
                review_entries.append(
                    make_review_item(
                        section="Metal Parts",
                        bom_row=idx + 2,
                        part_no=part_no,
                        part_name=part_name,
                        status="Review",
                        reason="하위 원재료 중 Qty 정보가 없는 항목이 있음",
                        check_item="원재료별 Qty, UOM",
                        action="Qty 누락 항목은 도면 확인 후 보완해 주세요",
                        reference=f"Child raw material={', '.join(child_part_nos)}",
                    )
                )

        else:
            review_entries.append(
                make_review_item(
                    section="Metal Parts",
                    bom_row=idx + 2,
                    part_no=part_no,
                    part_name=part_name,
                    status="Review",
                    reason="Metal 부품의 하위 원재료 정보가 없음",
                    check_item="하위 BOM 원재료 행, Qty, 금속 부품 상세 수량",
                    action="하위 원재료가 확인되면 BOM Qty를 기준으로 중량을 반영해 주세요",
                    reference=f"Detail row={textify(row.get('part_no'))}",
                )
            )

        candidate_rows.append({
            "Part No.": part_no,
            "Part Name": part_name,
            "Metal Type": metal_type,
            "thickness(mm)": thickness,
            "weight(g)": weight,
            "__source": source,
            "__has_weight": pd.notna(weight),
            "__bom_row": idx + 2,
            "__review_entries": review_entries,
        })

    if not candidate_rows:
        return pd.DataFrame(columns=METAL_COLUMNS), 0.0

    records_df = pd.DataFrame(candidate_rows)
    records_df["__source_rank"] = records_df["__source"].map(metal_source_rank).fillna(1).astype(int)
    records_df["__has_weight_sort"] = records_df["__has_weight"].astype(int)
    records_df["__weight_sort"] = pd.to_numeric(records_df["weight(g)"], errors="coerce").fillna(-1.0)
    records_df["__bom_row"] = pd.to_numeric(records_df["__bom_row"], errors="coerce").fillna(0).astype(int)

    records_df = records_df.sort_values(
        by=["Part No.", "__source_rank", "__has_weight_sort", "__weight_sort", "__bom_row"],
        ascending=[True, True, False, False, True],
        kind="mergesort",
    )

    records_df = records_df.drop_duplicates(subset=["Part No."], keep="first").copy()

    if review_items is not None:
        for _, row in records_df.iterrows():
            for review_entry in row.get("__review_entries", []) or []:
                review_items.append(review_entry)

    total_weight = pd.to_numeric(records_df["weight(g)"], errors="coerce").fillna(0.0).sum()
    return records_df[METAL_COLUMNS].reset_index(drop=True), float(total_weight)


def review_title(item):
    section = item.get("구분", "")
    reason = item.get("누락/미매칭 사유", "")
    mapping = {
        ("Plastic Parts", "BOM에 재질 정보가 없음"): "Plastic Parts - BOM 재질 정보 없음",
        ("Plastic Parts", "BOM에 재질 정보가 없음."): "Plastic Parts - BOM에 재질 정보가 없음.",
        ("Plastic Parts", "BOM에 재질 정보는 있으나 Resin List에 해당 원재료 정보가 없음"): "Plastic Parts - Resin List 미등록",
        ("Metal Parts", "BOM에 재질 정보가 없음"): "Metal Parts - BOM 재질 정보 없음",
        ("Metal Parts", "BOM에 재질 정보는 있으나 Metal List에 해당 원재료 정보가 없음"): "Metal Parts - Metal List 미등록",
        ("Metal Parts", "Metal 항목의 중량 정보가 없음"): "Metal Parts - 중량 정보 없음",
        ("Metal Parts", "하위 원재료는 있으나 Metal List에 해당 원재료 정보가 없음"): "Metal Parts - Metal List 미등록",
        ("Metal Parts", "하위 원재료 중 Qty 정보가 없는 항목이 있음"): "Metal Parts - 하위 원재료 Qty 누락",
        ("Metal Parts", "Metal 부품의 하위 원재료 정보가 없음"): "Metal Parts - 하위 원재료 정보 없음",
    }
    return mapping.get((section, reason), f"{section} - {reason}".strip(" -"))


def group_review_items(review_items):
    groups = []
    index_map = {}
    for item in review_items:
        title = review_title(item)
        if title not in index_map:
            index_map[title] = len(groups)
            groups.append({
                "title": title,
                "count": 0,
                "reason": item.get("누락/미매칭 사유", ""),
                "action": item.get("권장 조치", ""),
            })
        group = groups[index_map[title]]
        group["count"] += 1
    return groups


def render_review_table(review_groups):
    lines = []
    lines.append("| No. | 구분 | 검토 항목 | 건수 | 조치 |")
    lines.append("|---|---|---|---:|---|")
    if not review_groups:
        lines.append("| - | - | 검토 필요 항목 없음 | 0건 | - |")
        return lines

    for idx, group in enumerate(review_groups, start=1):
        title = group["title"]
        if " - " in title:
            section, issue = title.split(" - ", 1)
        else:
            section, issue = "", title
        count_text = f"{group['count']}건"
        lines.append(f"| {idx} | {section} | {issue} | {count_text} | {group['action']} |")
    return lines


def copy_row_format(ws, src_row, dest_row):
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dest_cell = ws.cell(row=dest_row, column=col)
        if src_cell.has_style:
            dest_cell.font = copy(src_cell.font)
            dest_cell.border = copy(src_cell.border)
            dest_cell.fill = copy(src_cell.fill)
            dest_cell.number_format = copy(src_cell.number_format)
            dest_cell.protection = copy(src_cell.protection)
            dest_cell.alignment = copy(src_cell.alignment)
    ws.row_dimensions[dest_row].height = ws.row_dimensions[src_row].height


def set_cell_value_safe(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col, value=value)
                return
        return
    cell.value = value


def clear_row(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        set_cell_value_safe(ws, row, col, None)


def find_target_sheet(wb):
    exact = "유첨.Plastic 부품 재질 List"
    if exact in wb.sheetnames:
        return wb[exact]
    for ws in wb.worksheets:
        if "Plastic" in ws.title or "재질" in ws.title:
            return ws
    raise ValueError("Target sheet not found: 유첨.Plastic 부품 재질 List")


def write_table(ws, start_row, template_end_row, records, columns, copy_source_row):
    template_rows = template_end_row - start_row + 1
    needed = len(records)
    added = max(0, needed - template_rows)

    if added > 0:
        ws.insert_rows(template_end_row, amount=added)
        for i in range(added):
            copy_row_format(ws, copy_source_row, template_end_row + i)

    for i, record in enumerate(records):
        row_num = start_row + i
        for col_num, field_name in enumerate(columns, start=3):
            set_cell_value_safe(ws, row_num, col_num, record.get(field_name, ""))

    if needed < template_rows:
        for row_num in range(start_row + needed, template_end_row + 1):
            clear_row(ws, row_num, 3, 2 + len(columns))

    return added


def build_workbook(bom_path, resin_source=DEFAULT_RESIN_SOURCE, metal_source=DEFAULT_METAL_SOURCE, template_source=DEFAULT_TEMPLATE_SOURCE, header_row=1, sheet_name=None, output_path=None, model_name=None):
    bom_df = load_bom(bom_path, header_row=header_row, sheet_name=sheet_name)
    resin_df = load_resin_reference(resin_source)
    metal_df = load_metal_reference(metal_source)

    review_items = []
    plastic_df, plastic_total_weight, plastic_total_pcr_weight = extract_plastic_records(bom_df, resin_df, review_items=review_items)
    metal_out_df, metal_total_weight = extract_metal_records(bom_df, metal_df, review_items=review_items)

    if model_name is None:
        stem = Path(bom_path).stem
        model_name = re.split(r"[_\-]", stem)[0]
    file_model_name = compact_model_name(model_name) or textify(model_name)

    wb = load_workbook(BytesIO(source_to_bytes(template_source)))
    ws = find_target_sheet(wb)
    ws["C11"] = model_name

    plastic_added = write_table(ws, 17, 19, plastic_df.to_dict("records"), PLASTIC_COLUMNS, copy_source_row=18)
    plastic_summary_row = 20 + plastic_added
    plastic_last_row = 16 + len(plastic_df)

    if len(plastic_df) == 0:
        ws["F20"] = 0
        ws["R20"] = 0
        ws["Q20"] = 0
    else:
        ws[f"F{plastic_summary_row}"] = f"=SUM(F17:F{plastic_last_row})"
        ws[f"R{plastic_summary_row}"] = f"=SUM(R17:R{plastic_last_row})"
        ws[f"Q{plastic_summary_row}"] = f"=IFERROR(R{plastic_summary_row}/F{plastic_summary_row},0)"

    metal_start = 33 + plastic_added
    metal_end = 35 + plastic_added
    write_table(ws, metal_start, metal_end, metal_out_df.to_dict("records"), METAL_COLUMNS, copy_source_row=metal_start + 1)


    if output_path is None:
        template_name = source_filename(template_source)
        output_path = Path(bom_path).with_name(f"{file_model_name}_{make_clickable_filename(template_name)}")
    else:
        output_path = Path(output_path)

    if output_path.exists():
        output_path = unique_file_path(output_path.parent, output_path.name)

    output_path = str(output_path)
    output_uri = Path(output_path).as_uri()

    wb.save(output_path)

    chat_output_dir = Path.cwd() / "generated_files"
    chat_output_dir.mkdir(parents=True, exist_ok=True)
    chat_output_name = make_clickable_filename(Path(output_path).name)
    chat_output_path = unique_file_path(chat_output_dir, chat_output_name)
    if Path(output_path).resolve() != chat_output_path.resolve():
        shutil.copy2(output_path, chat_output_path)
    display_output_dir = chat_output_dir
    display_output_path = str(chat_output_path)
    display_output_dir_uri = display_output_dir.as_uri()
    display_output_uri = chat_output_path.as_uri()

    review_groups = group_review_items(review_items)

    return {
        "model_name": model_name,
        "plastic_rows": len(plastic_df),
        "plastic_total_weight": plastic_total_weight,
        "plastic_total_pcr_weight": plastic_total_pcr_weight,
        "metal_rows": len(metal_out_df),
        "metal_total_weight": metal_total_weight,
        "review_rows": len(review_items),
        "review_items": review_items,
        "review_groups": review_groups,
        "output_path": output_path,
        "output_uri": output_uri,
        "display_output_dir": str(display_output_dir),
        "display_output_dir_uri": display_output_dir_uri,
        "display_output_path": display_output_path,
        "display_output_uri": display_output_uri,
    }



def run_generation(
    bom_file,
    resin_file,
    metal_file,
    template_file,
    header_row: int,
    sheet_name: str,
    model_name: str,
    output_dir: str,
):
    skill = get_skill()

    bom_path = save_uploaded_file(bom_file, "bom")
    resin_source = (
        save_uploaded_file(resin_file, "resin") if resin_file is not None else skill.DEFAULT_RESIN_SOURCE
    )
    metal_source = (
        save_uploaded_file(metal_file, "metal") if metal_file is not None else skill.DEFAULT_METAL_SOURCE
    )
    template_source = (
        save_uploaded_file(template_file, "template")
        if template_file is not None
        else skill.DEFAULT_TEMPLATE_SOURCE
    )

    summary = skill.build_workbook(
        bom_path=str(bom_path),
        resin_source=str(resin_source),
        metal_source=str(metal_source),
        template_source=str(template_source),
        header_row=header_row,
        sheet_name=sheet_name or None,
        model_name=model_name or None,
    )

    # 화면 표시용 리스트를 별도로 추출한다.
    bom_df = skill.load_bom(str(bom_path), header_row=header_row, sheet_name=sheet_name or None)
    resin_df = skill.load_resin_reference(str(resin_source))
    metal_df = skill.load_metal_reference(str(metal_source))

    plastic_records, _, _ = skill.extract_plastic_records(bom_df, resin_df, review_items=[])
    metal_records, _ = skill.extract_metal_records(bom_df, metal_df, review_items=[])
    summary["plastic_records"] = plastic_records.to_dict("records")
    summary["metal_records"] = metal_records.to_dict("records")

    if output_dir:
        target_dir = Path(output_dir).expanduser()
        if not target_dir.is_absolute():
            target_dir = (APP_DIR / target_dir).resolve()
        target_dir.mkdir(parents=True, exist_ok=True)

        source_path = Path(summary["output_path"])
        final_path = target_dir / source_path.name
        if source_path.resolve() != final_path.resolve():
            shutil.copy2(source_path, final_path)

        summary["output_path"] = str(final_path)
        summary["display_output_path"] = str(final_path)
        summary["display_output_dir"] = str(target_dir)
        summary["display_output_dir_uri"] = target_dir.as_uri()
        summary["output_uri"] = final_path.as_uri()
        summary["display_output_uri"] = final_path.as_uri()

    return summary


# -----------------------------------------------------------------------------
# UI 구성
# -----------------------------------------------------------------------------
def inject_css():
    st.markdown(
        """
        <style>
        :root {
            --bg-1: #f8fafc;
            --bg-2: #eef2ff;
            --panel: rgba(255, 255, 255, 0.82);
            --line: rgba(148, 163, 184, 0.22);
            --text: #0f172a;
            --muted: #64748b;
            --brand: #2563eb;
            --brand-2: #7c3aed;
            --brand-3: #0f766e;
        }
        html, body {
            background:
                radial-gradient(circle at top left, rgba(37,99,235,0.10), transparent 28%),
                radial-gradient(circle at top right, rgba(124,58,237,0.08), transparent 24%),
                linear-gradient(180deg, var(--bg-1), var(--bg-2));
        }
        .stApp {
            background: transparent;
            color: var(--text);
        }
        .block-container {
            padding-top: 1rem;
            padding-bottom: 2rem;
            max-width: 1240px;
        }
        section[data-testid="stSidebar"] {
            background: linear-gradient(180deg, rgba(255,255,255,0.96), rgba(248,250,252,0.95));
            border-right: 1px solid rgba(148,163,184,0.18);
        }
        .hero {
            border-radius: 24px;
            padding: 1.2rem 1.35rem;
            background:
                linear-gradient(135deg, rgba(15,118,110,0.95), rgba(37,99,235,0.94) 55%, rgba(124,58,237,0.94));
            border: 1px solid rgba(255,255,255,0.18);
            box-shadow: 0 18px 48px rgba(15, 23, 42, 0.16);
            margin-bottom: 1rem;
            color: white;
        }
        .hero h1 {
            margin: 0;
            font-size: 1.78rem;
            line-height: 1.2;
            color: white;
            letter-spacing: -0.04em;
        }
        .hero p {
            margin: 0.45rem 0 0 0;
            color: rgba(255,255,255,0.92);
            font-size: 0.95rem;
        }
        .soft-card {
            border: 1px solid var(--line);
            border-radius: 18px;
            padding: 1rem 1.05rem;
            background: var(--panel);
            backdrop-filter: blur(16px);
            box-shadow: 0 16px 40px rgba(15, 23, 42, 0.06);
        }
        .section-title {
            font-size: 1rem;
            font-weight: 800;
            margin-bottom: 0.35rem;
            letter-spacing: -0.02em;
            color: var(--text);
        }
        .subtle {
            color: var(--muted);
            font-size: 0.9rem;
        }
        .kpi-strip {
            display: flex;
            gap: 10px;
            overflow-x: auto;
            padding: 2px 2px 8px 2px;
            margin: 8px 0 12px 0;
            scrollbar-width: thin;
        }
        .kpi-strip::-webkit-scrollbar {
            height: 8px;
        }
        .kpi-strip::-webkit-scrollbar-thumb {
            background: rgba(148,163,184,0.45);
            border-radius: 999px;
        }
        .kpi-card {
            flex: 0 0 170px;
            position: relative;
            overflow: hidden;
            border-radius: 16px;
            padding: 12px 13px 11px;
            background: rgba(255,255,255,0.88);
            border: 1px solid rgba(148,163,184,0.20);
            box-shadow: 0 12px 30px rgba(15, 23, 42, 0.05);
        }
        .kpi-card::before {
            content: "";
            position: absolute;
            inset: 0 auto auto 0;
            width: 100%;
            height: 4px;
            background: linear-gradient(90deg, var(--brand), var(--brand-2), var(--brand-3));
        }
        .kpi-label {
            font-size: 11px;
            color: var(--muted);
            font-weight: 700;
            letter-spacing: 0.02em;
        }
        .kpi-value {
            margin-top: 6px;
            font-size: 20px;
            font-weight: 900;
            color: var(--text);
            letter-spacing: -0.04em;
            line-height: 1.05;
            word-break: break-word;
        }
        .kpi-note { display: none; }
        .panel-shell {
            padding: 16px;
            border-radius: 22px;
            background: rgba(255,255,255,0.58);
            border: 1px solid rgba(148,163,184,0.18);
            backdrop-filter: blur(14px);
            box-shadow: 0 18px 42px rgba(15, 23, 42, 0.06);
        }
        .path-card {
            padding: 14px 15px;
            border-radius: 16px;
            background: rgba(248,250,252,0.92);
            border: 1px solid rgba(148,163,184,0.18);
            word-break: break-all;
        }
        .path-label {
            font-size: 12px;
            color: var(--muted);
            font-weight: 700;
            margin-bottom: 6px;
        }
        .path-value {
            font-size: 13px;
            color: var(--text);
            line-height: 1.5;
        }
        .download-hero {
            padding: 16px;
            border-radius: 18px;
            background: linear-gradient(135deg, rgba(37,99,235,0.10), rgba(124,58,237,0.10), rgba(15,118,110,0.10));
            border: 1px solid rgba(37,99,235,0.14);
        }
        .download-hero .section-title {
            margin-bottom: 0.15rem;
        }
        div[data-testid="stDownloadButton"] > button {
            font-weight: 900 !important;
            font-size: 1.28rem !important;
            padding-top: 1.25rem !important;
            padding-bottom: 1.25rem !important;
            min-height: 64px !important;
            border-radius: 16px !important;
            border: 0 !important;
            color: white !important;
            background: linear-gradient(135deg, var(--brand), #4f46e5) !important;
            box-shadow: 0 22px 44px rgba(37, 99, 235, 0.34) !important;
            letter-spacing: -0.01em;
        }
        div[data-testid="stDownloadButton"] > button:hover {
            transform: translateY(-1px);
            box-shadow: 0 26px 48px rgba(37, 99, 235, 0.42) !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_hero(lang: str):
    st.markdown(
        """
        <div class="hero">
            <h1>%s</h1>
            <p>
                %s
            </p>
        </div>
        """
        % (html.escape(tr(lang, "hero_title")), html.escape(tr(lang, "hero_desc"))),
        unsafe_allow_html=True,
    )


def render_sidebar():
    st.sidebar.markdown(f"**🌐 {tr('ko', 'language')}**")
    language = st.sidebar.radio(
        label="Language",
        options=["Korean", "English"],
        index=0,
        horizontal=True,
        label_visibility="collapsed",
    )
    lang = "en" if language == "English" else "ko"

    st.sidebar.title(tr(lang, "sidebar_title"))

    bom_file = st.sidebar.file_uploader(
        tr(lang, "bom_upload"),
        type=["xlsx", "xlsm"],
        help=(
            "Upload the BOM Excel file to generate the eco checklist."
            if lang == "en"
            else "친환경 체크리스트를 만들 BOM Excel 파일을 업로드하세요."
        ),
    )

    suggested_model_name = infer_model_name(bom_file) if bom_file is not None else ""
    if bom_file is not None:
        if suggested_model_name:
            st.sidebar.success(
                f"{tr(lang, 'model_name')}: {suggested_model_name}"
            )
        else:
            st.sidebar.warning(
                "Could not detect a model name. Please review the file manually."
                if lang == "en"
                else "자동 감지 모델명을 찾지 못했습니다. 직접 입력해 주세요."
            )

    auto_generate = st.sidebar.checkbox(tr(lang, "auto_generate"), value=True)

    st.sidebar.caption(tr(lang, "sidebar_tip"))

    sidebar_notice = st.sidebar.empty()

    return {
        "lang": lang,
        "bom_file": bom_file,
        "resin_file": None,
        "metal_file": None,
        "template_file": None,
        "header_row": 1,
        "sheet_name": "",
        "model_name": "",
        "output_dir": str(APP_DIR / "generated_files"),
        "suggested_model_name": suggested_model_name.strip(),
        "auto_generate": auto_generate,
        "sidebar_notice": sidebar_notice,
    }


def review_table_markdown(review_groups, lang: str):
    lines = [
        "| No. | Category | Review Item | Count | Action |"
        if lang == "en"
        else "| No. | 구분 | 검토 항목 | 건수 | 조치 |",
        "|---|---|---|---:|---|",
    ]
    if not review_groups:
        lines.append(
            "| - | - | No review items | 0ea | - |"
            if lang == "en"
            else "| - | - | 검토 필요 항목 없음 | 0건 | - |"
        )
        return "\n".join(lines)

    for idx, item in enumerate(review_groups, start=1):
        title = item.get("title", "")
        if " - " in title:
            section, issue = title.split(" - ", 1)
        else:
            section, issue = "", title
        count = f"{item.get('count', 0)}{tr(lang, 'count_suffix')}"
        action = item.get("action", "-") or "-"
        action = str(action).replace("\n", " ")
        lines.append(f"| {idx} | {section} | {issue} | {count} | {action} |")
    return "\n".join(lines)


def render_summary(summary, bom_file_name: str, lang: str):
    plastic_total_weight = summary.get("plastic_total_weight", 0.0)
    plastic_pcr_weight = summary.get("plastic_total_pcr_weight", 0.0)
    plastic_pcr_ratio = (
        (float(plastic_pcr_weight) / float(plastic_total_weight) * 100.0)
        if float(plastic_total_weight or 0) > 0
        else 0.0
    )

    st.markdown(
        f"""
        <div class="kpi-strip">
            <div class="kpi-card">
                <div class="kpi-label">{html.escape(tr(lang, "model_name"))}</div>
                <div class="kpi-value">{html.escape(str(summary.get("model_name", "-")))}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">{html.escape(tr(lang, "plastic_total_weight"))}</div>
                <div class="kpi-value">{html.escape(format_grams(plastic_total_weight))}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">{html.escape(tr(lang, "pcr_weight"))}</div>
                <div class="kpi-value">{html.escape(format_grams(plastic_pcr_weight))}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">{html.escape(tr(lang, "pcr_ratio"))}</div>
                <div class="kpi-value">{html.escape(format_percent(plastic_pcr_ratio))}</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(f"### {tr(lang, 'part_list')}")
    plastic_records = summary.get("plastic_records", [])
    metal_records = summary.get("metal_records", [])

    part_tabs = st.tabs([tr(lang, "plastic_part_list"), tr(lang, "metal_part_list")])
    with part_tabs[0]:
        st.caption(f"{tr(lang, 'plastic_rows')}: {len(plastic_records)}")
        if plastic_records:
            st.dataframe(
                pd.DataFrame(plastic_records),
                use_container_width=True,
                hide_index=True,
                height=420,
            )
        else:
            st.info("No Plastic Part List to display." if lang == "en" else "표시할 Plastic Part List가 없습니다.")

    with part_tabs[1]:
        st.caption(f"{tr(lang, 'metal_rows')}: {len(metal_records)}")
        if metal_records:
            st.dataframe(
                pd.DataFrame(metal_records),
                use_container_width=True,
                hide_index=True,
                height=360,
            )
        else:
            st.info("No Metal Part List to display." if lang == "en" else "표시할 Metal Part List가 없습니다.")

    st.markdown(f"### {tr(lang, 'summary_title')}")
    result_col, info_col = st.columns([1.3, 1])

    output_path = Path(summary["output_path"])
    output_exists = output_path.exists()

    with result_col:
        st.markdown(
            """
            <div class="soft-card download-hero">
                <div class="section-title">%s</div>
                <div class="subtle">%s</div>
            </div>
            """
            % (html.escape(tr(lang, "download")), html.escape(tr(lang, "generated_caption"))),
            unsafe_allow_html=True,
        )
        if output_exists:
            st.download_button(
                label=tr(lang, "download_button"),
                data=output_path.read_bytes(),
                file_name=output_path.name,
                mime=MIME_XLSX,
                use_container_width=True,
            )
        else:
            st.error("Generated file not found. Please run generation again." if lang == "en" else "생성된 파일을 찾을 수 없습니다. 다시 생성해 주세요.")

    with info_col:
        st.markdown(
            f"""
            <div class="panel-shell">
                <div class="section-title">{html.escape(tr(lang, "input_output"))}</div>
                <div class="path-card" style="margin-top:10px;">
                    <div class="path-label">{html.escape(tr(lang, "input_file"))}</div>
                    <div class="path-value">{html.escape(str(bom_file_name))}</div>
                </div>
                <div class="path-card" style="margin-top:10px;">
                    <div class="path-label">{html.escape(tr(lang, "output_file"))}</div>
                    <div class="path-value">{html.escape(str(summary.get("display_output_path", "-")))}</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown(f"### {tr(lang, 'review_title')}")
    st.markdown(review_table_markdown(summary.get("review_groups", []), lang))

    with st.expander(tr(lang, "review_details"), expanded=False):
        review_items = summary.get("review_items", [])
        if not review_items:
            st.info("No review detail items." if lang == "en" else "검토 상세 항목이 없습니다.")
        else:
            st.json(review_items)


def render_empty_state(lang: str):
    st.info(tr(lang, "input_file_wait"))
    st.markdown(
        """
        <div class="soft-card">
            <div class="section-title">%s</div>
            <ul>
                <li>%s</li>
                <li>%s</li>
                <li>%s</li>
            </ul>
        </div>
        """
        % (
            html.escape(tr(lang, "support_title")),
            html.escape(tr(lang, "support_1")),
            html.escape(tr(lang, "support_2")),
            html.escape(tr(lang, "support_3")),
        ),
        unsafe_allow_html=True,
    )


def main():
    st.set_page_config(
        page_title="BOM Eco Checklist",
        page_icon="♻️",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    inject_css()
    inputs = render_sidebar()
    lang = inputs["lang"]
    render_hero(lang)
    st.write("")
    bom_file = inputs["bom_file"]
    sidebar_notice = inputs["sidebar_notice"]

    if bom_file is None:
        render_empty_state(lang)
        return

    effective_model_name = (
        inputs["model_name"].strip()
        or inputs["suggested_model_name"].strip()
        or infer_model_name(bom_file)
    )

    signature = build_signature(
        bom_file=bom_file,
        resin_file=inputs["resin_file"],
        metal_file=inputs["metal_file"],
        template_file=inputs["template_file"],
        header_row=inputs["header_row"],
        sheet_name=inputs["sheet_name"],
        model_name=effective_model_name,
        output_dir=inputs["output_dir"],
        auto_generate=inputs["auto_generate"],
    )

    manual_generate_clicked = False
    if not inputs["auto_generate"]:
        st.info("Auto generation is off. Click the button below to generate." if lang == "en" else "자동 생성이 꺼져 있습니다. 아래 버튼을 눌러 생성해 주세요.")
        manual_generate_clicked = st.button(
            "Generate Eco Checklist" if lang == "en" else "친환경 체크리스트 생성",
            type="primary",
            use_container_width=True,
        )

    should_process = inputs["auto_generate"] or manual_generate_clicked
    cached_summary = st.session_state.get("last_summary")

    if should_process:
        if st.session_state.get("last_signature") != signature:
            with st.spinner("Analyzing BOM and generating the eco checklist..." if lang == "en" else "BOM을 분석하고 친환경 체크리스트를 생성하는 중입니다..."):
                try:
                    summary = run_generation(
                        bom_file=bom_file,
                        resin_file=inputs["resin_file"],
                        metal_file=inputs["metal_file"],
                        template_file=inputs["template_file"],
                        header_row=inputs["header_row"],
                        sheet_name=inputs["sheet_name"],
                        model_name=effective_model_name,
                        output_dir=inputs["output_dir"],
                    )
                    st.session_state["last_signature"] = signature
                    st.session_state["last_summary"] = summary
                    st.session_state["last_bom_name"] = bom_file.name
                    sidebar_notice.success(tr(lang, "generated_message"))
                except Exception as exc:
                    st.error(f"Generation failed: {exc}" if lang == "en" else f"생성 중 오류가 발생했습니다: {exc}")
                    with st.expander("Error details" if lang == "en" else "오류 상세", expanded=True):
                        st.code(traceback.format_exc())
                    return

        summary = st.session_state.get("last_summary")
        if summary:
            render_summary(summary, st.session_state.get("last_bom_name", bom_file.name), lang)
        else:
            st.warning("Could not load the result. Please try again." if lang == "en" else "결과를 불러오지 못했습니다. 다시 시도해 주세요.")
    elif cached_summary and st.session_state.get("last_signature") == signature:
        render_summary(cached_summary, st.session_state.get("last_bom_name", bom_file.name), lang)
        sidebar_notice.success(tr(lang, "generated_message"))
    elif not inputs["auto_generate"]:
        st.caption("The file is uploaded but generation has not run yet." if lang == "en" else "파일은 업로드되었지만 아직 생성이 실행되지 않았습니다.")

    st.markdown("---")
    st.caption(
        f"{'Base logic' if lang == 'en' else '기준 모듈'}: {SKILL_SCRIPT} | {'Output location' if lang == 'en' else '출력 위치'}: {APP_DIR / 'generated_files'}"
    )


if __name__ == "__main__":
    main()
